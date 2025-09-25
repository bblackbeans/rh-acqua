from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Avg, Count, Q
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import JsonResponse, HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from users.models import UserProfile, TechnicalSkill, SoftSkill, Certification, Language, Education as UserEducation, Experience as UserExperience
from vacancies.models import Vacancy, Hospital
from .models import Application, ApplicationEvaluation, Resume, Education, WorkExperience, ApplicationFavorite, ApplicationComplementaryInfo
from .forms import (
    ApplicationForm, ApplicationEvaluationForm, ResumeForm, 
    EducationForm, WorkExperienceForm, ApplicationStatusForm
)
from .serializers import (
    ApplicationSerializer, ApplicationCreateSerializer, ApplicationStatusUpdateSerializer,
    ApplicationEvaluationSerializer, ResumeSerializer, ResumeCreateUpdateSerializer,
    EducationSerializer, EducationCreateUpdateSerializer,
    WorkExperienceSerializer, WorkExperienceCreateUpdateSerializer
)
from .permissions import IsOwnerOrRecruiter, IsRecruiterOrAdmin, IsResumeOwner, IsEducationOrExperienceOwner

import csv
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta


# Views para interface web

@login_required
def application_list(request):
    """
    Exibe a lista de candidaturas do usuário atual ou todas as candidaturas para recrutadores.
    """
    user_profile = request.user.profile
    
    # Inicializar variáveis de filtro para todos os usuários
    status_filter = request.GET.get('status')
    vacancy_filter = request.GET.get('vacancy')
    hospital_filter = request.GET.get('hospital')
    date_filter = request.GET.get('date')
    search_query = request.GET.get('search')
    score_sort = request.GET.get('score_sort')
    favorites_filter = request.GET.get('favorites')
    
    if request.user.role == 'candidate':
        # Candidatos veem apenas suas próprias candidaturas
        applications = Application.objects.filter(candidate=user_profile)
        template = 'applications/candidate_application_list.html'
    else:
        # Recrutadores e administradores veem todas as candidaturas
        applications = Application.objects.all()
        
        if status_filter:
            applications = applications.filter(status=status_filter)
        
        if vacancy_filter:
            applications = applications.filter(vacancy_id=vacancy_filter)
            
        if hospital_filter:
            applications = applications.filter(vacancy__hospital_id=hospital_filter)
            
        if date_filter:
            if date_filter == '7':
                applications = applications.filter(created_at__gte=timezone.now() - timezone.timedelta(days=7))
            elif date_filter == '30':
                applications = applications.filter(created_at__gte=timezone.now() - timezone.timedelta(days=30))
            elif date_filter == '90':
                applications = applications.filter(created_at__gte=timezone.now() - timezone.timedelta(days=90))
                
        if search_query:
            applications = applications.filter(
                Q(candidate__user__first_name__icontains=search_query) |
                Q(candidate__user__last_name__icontains=search_query) |
                Q(candidate__user__email__icontains=search_query) |
                Q(vacancy__title__icontains=search_query)
            )
            
        if favorites_filter:
            # Filtro para favoritos
            if favorites_filter == 'true':
                applications = applications.filter(favorites__recruiter=user_profile)
            
        # Ordenação por score
        if score_sort == 'high_to_low':
            applications = applications.annotate(
                avg_score=(Avg('evaluations__technical_score') + 
                         Avg('evaluations__experience_score') + 
                         Avg('evaluations__cultural_fit_score')) / 3
            ).order_by('-avg_score')
        elif score_sort == 'low_to_high':
            applications = applications.annotate(
                avg_score=(Avg('evaluations__technical_score') + 
                         Avg('evaluations__experience_score') + 
                         Avg('evaluations__cultural_fit_score')) / 3
            ).order_by('avg_score')
        else:
            # Ordenação padrão por data de criação
            applications = applications.order_by('-created_at')
            
        template = 'applications/recruiter_application_list.html'
    
    # Contagem de candidaturas por status
    status_counts = {}
    if request.user.role in ['recruiter', 'recrutador', 'admin']:
        # Adiciona informações sobre favoritos e score médio
        for application in applications:
            application.is_favorite = application.favorites.filter(recruiter=user_profile).exists()
            
            # Calcula o score médio se houver avaliações
            if application.evaluations.exists():
                evaluation = application.evaluations.first()
                total_score = evaluation.technical_score + evaluation.experience_score + evaluation.cultural_fit_score
                application.avg_score = round(total_score / 3, 1)
            else:
                application.avg_score = None
        
        status_counts = {
            'total': applications.count(),
            'pending': applications.filter(status='pending').count(),
            'under_review': applications.filter(status='under_review').count(),
            'interview': applications.filter(status='interview').count(),
            'approved': applications.filter(status='approved').count(),
            'rejected': applications.filter(status='rejected').count(),
            'withdrawn': applications.filter(status='withdrawn').count(),
        }
    
    context = {
        'applications': applications,
        'status_counts': status_counts,
        'status_filter': status_filter,
        'vacancy_filter': vacancy_filter,
        'hospital_filter': hospital_filter,
        'date_filter': date_filter,
        'search_query': search_query,
        'score_sort': score_sort,
        'favorites_filter': favorites_filter,
        'page_title': _('Candidaturas'),
    }
    
    return render(request, template, context)


@login_required
def application_detail(request, pk):
    """
    Exibe os detalhes de uma candidatura específica.
    """
    application = get_object_or_404(Application, pk=pk)
    
    # Verifica se o usuário tem perfil, se não tiver, cria um
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=request.user)
    
    # Verifica permissões
    if request.user.role == 'candidate' and application.candidate != user_profile:
        messages.error(request, _('Você não tem permissão para acessar esta candidatura.'))
        return redirect('applications:application_list')
    
    # Verifica se é favorito
    application.is_favorite = False
    if request.user.role in ['recruiter', 'recrutador', 'admin']:
        application.is_favorite = application.favorites.filter(recruiter=user_profile).exists()
    
    # Busca todas as informações do candidato
    candidate = application.candidate
    
    # Busca currículo detalhado se existir
    try:
        detailed_resume = Resume.objects.get(candidate=candidate)
    except Resume.DoesNotExist:
        detailed_resume = None
    
    # Busca informações complementares se existir
    try:
        complementary_info = ApplicationComplementaryInfo.objects.get(application=application)
    except ApplicationComplementaryInfo.DoesNotExist:
        complementary_info = None
    
    # Formulário de avaliação para recrutadores
    evaluation_form = None
    existing_evaluation = None
    
    if request.user.role in ['recruiter', 'recrutador', 'admin']:
        # Verifica se já existe uma avaliação deste recrutador
        try:
            existing_evaluation = ApplicationEvaluation.objects.get(
                application=application, 
                evaluator=user_profile
            )
        except ApplicationEvaluation.DoesNotExist:
            existing_evaluation = None
        
        # Processa o formulário único de salvar todas as informações
        if request.method == 'POST' and 'save_all' in request.POST:
            # Atualiza status se fornecido
            if 'status' in request.POST:
                application.status = request.POST['status']
                application.save()
            
            # Atualiza notas do recrutador
            if 'recruiter_notes' in request.POST:
                application.recruiter_notes = request.POST.get('recruiter_notes', '')
                application.save()
            
            # Processa favoritar
            is_favorite = request.POST.get('is_favorite') == 'on'
            if is_favorite:
                ApplicationFavorite.objects.get_or_create(
                    application=application,
                    recruiter=user_profile
                )
            else:
                ApplicationFavorite.objects.filter(
                    application=application,
                    recruiter=user_profile
                ).delete()
            
            # Processa avaliação
            if existing_evaluation:
                # Atualiza avaliação existente
                existing_evaluation.technical_score = request.POST.get('technical_score', 0)
                existing_evaluation.experience_score = request.POST.get('experience_score', 0)
                existing_evaluation.cultural_fit_score = request.POST.get('cultural_fit_score', 0)
                existing_evaluation.comments = request.POST.get('comments', '')
                existing_evaluation.save()
            else:
                # Cria nova avaliação
                evaluation = ApplicationEvaluation.objects.create(
                    application=application,
                    evaluator=user_profile,
                    technical_score=request.POST.get('technical_score', 0),
                    experience_score=request.POST.get('experience_score', 0),
                    cultural_fit_score=request.POST.get('cultural_fit_score', 0),
                    comments=request.POST.get('comments', '')
                )
            
            messages.success(request, _('Todas as informações foram salvas com sucesso!'))
            return redirect('applications:application_detail', pk=pk)
        
        # Inicializa formulários com dados existentes
        if existing_evaluation:
            evaluation_form = ApplicationEvaluationForm(instance=existing_evaluation)
        else:
            evaluation_form = ApplicationEvaluationForm()
    
    # Formulário de atualização de status para recrutadores
    status_form = None
    if request.user.role in ['recruiter', 'recrutador', 'admin']:
        status_form = ApplicationStatusForm(instance=application)
    
    context = {
        'application': application,
        'candidate': candidate,
        'detailed_resume': detailed_resume,
        'complementary_info': complementary_info,
        'evaluation_form': evaluation_form,
        'status_form': status_form,
        'existing_evaluation': existing_evaluation,
        'evaluations': application.evaluations.all(),
        'page_title': _('Detalhes da Candidatura'),
    }
    
    return render(request, 'applications/application_detail.html', context)


@login_required
def apply_for_vacancy(request, vacancy_id):
    """
    Permite que um candidato se candidate a uma vaga.
    """
    user_profile = request.user.profile
    
    # Verifica se o usuário é um candidato
    if request.user.role != 'candidate':
        messages.error(request, _('Apenas candidatos podem se candidatar a vagas.'))
        return redirect('vacancy_list')
    
    vacancy = get_object_or_404(Vacancy, pk=vacancy_id)
    
    # Verifica se a vaga está aberta para candidaturas
    if vacancy.status != 'open':
        messages.error(request, _('Esta vaga não está aberta para candidaturas.'))
        return redirect('vacancy_detail', pk=vacancy_id)
    
    # Verifica se o candidato já se candidatou a esta vaga
    if Application.objects.filter(candidate=user_profile, vacancy=vacancy).exists():
        messages.error(request, _('Você já se candidatou a esta vaga.'))
        return redirect('vacancy_detail', pk=vacancy_id)
    
    if request.method == 'POST':
        form = ApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.candidate = user_profile
            application.vacancy = vacancy
            application.save()
            messages.success(request, _('Candidatura enviada com sucesso!'))
            return redirect('application_list')
    else:
        form = ApplicationForm()
    
    context = {
        'form': form,
        'vacancy': vacancy,
        'page_title': _('Candidatar-se à Vaga'),
    }
    
    return render(request, 'applications/apply_form.html', context)


@login_required
def resume_edit(request):
    """
    Permite que um candidato edite seu currículo detalhado.
    """
    user_profile = request.user.profile
    
    # Verifica se o usuário é um candidato
    if request.user.role != 'candidate':
        messages.error(request, _('Apenas candidatos podem editar currículos.'))
        return redirect('dashboard')
    
    # Obtém ou cria o currículo do candidato
    resume, created = Resume.objects.get_or_create(candidate=user_profile)
    
    if request.method == 'POST':
        form = ResumeForm(request.POST, instance=resume)
        if form.is_valid():
            form.save()
            messages.success(request, _('Currículo atualizado com sucesso!'))
            return redirect('resume_detail')
    else:
        form = ResumeForm(instance=resume)
    
    context = {
        'form': form,
        'resume': resume,
        'page_title': _('Editar Currículo'),
    }
    
    return render(request, 'applications/resume_edit.html', context)


@login_required
def resume_detail(request):
    """
    Exibe as informações completas da candidatura do candidato.
    """
    # Verifica se o usuário tem perfil, se não tiver, cria um
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=request.user)
    
    # Verifica se é um recrutador vendo candidatura de outro candidato
    candidate_id = request.GET.get('candidate_id')
    
    if candidate_id and request.user.role in ['recruiter', 'recrutador', 'admin']:
        # Recrutador vendo candidatura de outro candidato
        candidate_profile = get_object_or_404(UserProfile, id=candidate_id)
        
        # Busca a candidatura mais recente deste candidato
        try:
            application = Application.objects.filter(candidate=candidate_profile).latest('created_at')
        except Application.DoesNotExist:
            messages.error(request, _('Nenhuma candidatura encontrada para este candidato.'))
            return redirect('applications:candidaturas')
        
        page_title = f'Informações da Candidatura - {candidate_profile.user.get_full_name()}'
        
        # Busca informações complementares da candidatura
        try:
            complementary_info = ApplicationComplementaryInfo.objects.get(application=application)
        except ApplicationComplementaryInfo.DoesNotExist:
            complementary_info = None
            
    else:
        # Candidato vendo sua própria candidatura
        if request.user.role != 'candidate':
            messages.error(request, _('Apenas candidatos podem ver suas próprias candidaturas.'))
            return redirect('users:login')
        
        # Busca a candidatura mais recente do candidato
        try:
            application = Application.objects.filter(candidate=user_profile).latest('created_at')
        except Application.DoesNotExist:
            messages.error(request, _('Nenhuma candidatura encontrada.'))
            return redirect('applications:minhas_candidaturas')
        
        page_title = _('Minha Candidatura')
        
        # Busca informações complementares da candidatura
        try:
            complementary_info = ApplicationComplementaryInfo.objects.get(application=application)
        except ApplicationComplementaryInfo.DoesNotExist:
            complementary_info = None
    
    # Busca currículo se existir
    try:
        resume = Resume.objects.get(candidate=candidate_profile if candidate_id else user_profile)
        education_list = resume.education.all()
        experience_list = resume.work_experiences.all()
    except Resume.DoesNotExist:
        resume = None
        education_list = []
        experience_list = []
    
    # Busca informações do usuário
    candidate_user = candidate_profile.user if candidate_id else user_profile.user
    
    # Busca dados do currículo do usuário
    technical_skills = candidate_user.technical_skills.all()
    soft_skills = candidate_user.soft_skills.all()
    certifications = candidate_user.certifications.all()
    languages = candidate_user.languages.all()
    user_educations = candidate_user.educations.all()
    user_experiences = candidate_user.experiences.all()
    
    context = {
        'application': application,
        'candidate_profile': candidate_profile if candidate_id else user_profile,
        'complementary_info': complementary_info,
        'resume': resume,
        'education_list': education_list,  # Do Resume antigo
        'experience_list': experience_list,  # Do Resume antigo
        'user_educations': user_educations,  # Do User (dados atuais)
        'user_experiences': user_experiences,  # Do User (dados atuais)
        'technical_skills': technical_skills,
        'soft_skills': soft_skills,
        'certifications': certifications,
        'languages': languages,
        'page_title': page_title,
    }
    
    return render(request, 'applications/resume_detail.html', context)


@login_required
def education_create(request):
    """
    Permite que um candidato adicione uma formação educacional ao seu currículo.
    """
    user_profile = request.user.profile
    
    # Verifica se o usuário é um candidato
    if request.user.role != 'candidate':
        messages.error(request, _('Apenas candidatos podem adicionar formação educacional.'))
        return redirect('dashboard')
    
    # Obtém o currículo do candidato
    resume = get_object_or_404(Resume, candidate=user_profile)
    
    if request.method == 'POST':
        form = EducationForm(request.POST)
        if form.is_valid():
            education = form.save(commit=False)
            education.resume = resume
            education.save()
            messages.success(request, _('Formação educacional adicionada com sucesso!'))
            return redirect('resume_detail')
    else:
        form = EducationForm()
    
    context = {
        'form': form,
        'page_title': _('Adicionar Formação Educacional'),
    }
    
    return render(request, 'applications/education_form.html', context)


@login_required
def education_edit(request, pk):
    """
    Permite que um candidato edite uma formação educacional.
    """
    user_profile = request.user.profile
    
    # Obtém a formação educacional
    education = get_object_or_404(Education, pk=pk)
    
    # Verifica se o usuário é o proprietário do currículo
    if education.resume.candidate != user_profile:
        messages.error(request, _('Você não tem permissão para editar esta formação educacional.'))
        return redirect('resume_detail')
    
    if request.method == 'POST':
        form = EducationForm(request.POST, instance=education)
        if form.is_valid():
            form.save()
            messages.success(request, _('Formação educacional atualizada com sucesso!'))
            return redirect('resume_detail')
    else:
        form = EducationForm(instance=education)
    
    context = {
        'form': form,
        'education': education,
        'page_title': _('Editar Formação Educacional'),
    }
    
    return render(request, 'applications/education_form.html', context)


@login_required
def education_delete(request, pk):
    """
    Permite que um candidato exclua uma formação educacional.
    """
    user_profile = request.user.profile
    
    # Obtém a formação educacional
    education = get_object_or_404(Education, pk=pk)
    
    # Verifica se o usuário é o proprietário do currículo
    if education.resume.candidate != user_profile:
        messages.error(request, _('Você não tem permissão para excluir esta formação educacional.'))
        return redirect('resume_detail')
    
    if request.method == 'POST':
        education.delete()
        messages.success(request, _('Formação educacional excluída com sucesso!'))
        return redirect('resume_detail')
    
    context = {
        'education': education,
        'page_title': _('Excluir Formação Educacional'),
    }
    
    return render(request, 'applications/education_confirm_delete.html', context)


@login_required
def work_experience_create(request):
    """
    Permite que um candidato adicione uma experiência profissional ao seu currículo.
    """
    user_profile = request.user.profile
    
    # Verifica se o usuário é um candidato
    if request.user.role != 'candidate':
        messages.error(request, _('Apenas candidatos podem adicionar experiência profissional.'))
        return redirect('dashboard')
    
    # Obtém o currículo do candidato
    resume = get_object_or_404(Resume, candidate=user_profile)
    
    if request.method == 'POST':
        form = WorkExperienceForm(request.POST)
        if form.is_valid():
            experience = form.save(commit=False)
            experience.resume = resume
            experience.save()
            messages.success(request, _('Experiência profissional adicionada com sucesso!'))
            return redirect('resume_detail')
    else:
        form = WorkExperienceForm()
    
    context = {
        'form': form,
        'page_title': _('Adicionar Experiência Profissional'),
    }
    
    return render(request, 'applications/work_experience_form.html', context)


@login_required
def work_experience_edit(request, pk):
    """
    Permite que um candidato edite uma experiência profissional.
    """
    user_profile = request.user.profile
    
    # Obtém a experiência profissional
    experience = get_object_or_404(WorkExperience, pk=pk)
    
    # Verifica se o usuário é o proprietário do currículo
    if experience.resume.candidate != user_profile:
        messages.error(request, _('Você não tem permissão para editar esta experiência profissional.'))
        return redirect('resume_detail')
    
    if request.method == 'POST':
        form = WorkExperienceForm(request.POST, instance=experience)
        if form.is_valid():
            form.save()
            messages.success(request, _('Experiência profissional atualizada com sucesso!'))
            return redirect('resume_detail')
    else:
        form = WorkExperienceForm(instance=experience)
    
    context = {
        'form': form,
        'experience': experience,
        'page_title': _('Editar Experiência Profissional'),
    }
    
    return render(request, 'applications/work_experience_form.html', context)


@login_required
def work_experience_delete(request, pk):
    """
    Permite que um candidato exclua uma experiência profissional.
    """
    user_profile = request.user.profile
    
    # Obtém a experiência profissional
    experience = get_object_or_404(WorkExperience, pk=pk)
    
    # Verifica se o usuário é o proprietário do currículo
    if experience.resume.candidate != user_profile:
        messages.error(request, _('Você não tem permissão para excluir esta experiência profissional.'))
        return redirect('resume_detail')
    
    if request.method == 'POST':
        experience.delete()
        messages.success(request, _('Experiência profissional excluída com sucesso!'))
        return redirect('resume_detail')
    
    context = {
        'experience': experience,
        'page_title': _('Excluir Experiência Profissional'),
    }
    
    return render(request, 'applications/work_experience_confirm_delete.html', context)


@login_required
def candidaturas(request):
    """
    Exibe a página de candidaturas para recrutadores.
    """
    # Verifica se o usuário é um recrutador
    if request.user.role not in ['recruiter', 'recrutador', 'admin']:
        messages.error(request, _('Apenas recrutadores podem acessar esta página.'))
        return redirect('home')
    
    user_profile = request.user.profile
    
    # Obtém todas as candidaturas
    applications = Application.objects.all()
    
    # Filtros
    status_filter = request.GET.get('status', '')
    if status_filter:
        applications = applications.filter(status=status_filter)
    
    vacancy_filter = request.GET.get('vacancy', '')
    if vacancy_filter:
        applications = applications.filter(vacancy_id=vacancy_filter)
    
    hospital_filter = request.GET.get('hospital', '')
    if hospital_filter:
        applications = applications.filter(vacancy__hospital_id=hospital_filter)
    
    date_filter = request.GET.get('date', '')
    if date_filter:
        from django.utils import timezone
        from datetime import timedelta
        
        today = timezone.now().date()
        if date_filter == 'today':
            applications = applications.filter(created_at__date=today)
        elif date_filter == 'week':
            week_ago = today - timedelta(days=7)
            applications = applications.filter(created_at__date__gte=week_ago)
        elif date_filter == 'month':
            month_ago = today - timedelta(days=30)
            applications = applications.filter(created_at__date__gte=month_ago)
    
    search_query = request.GET.get('search', '')
    if search_query:
        applications = applications.filter(
            Q(candidate__user__first_name__icontains=search_query) |
            Q(candidate__user__last_name__icontains=search_query) |
            Q(candidate__user__email__icontains=search_query)
        )
    
    # Filtro de favoritos
    favorites_filter = request.GET.get('favorites', '')
    if favorites_filter == 'true':
        applications = applications.filter(favorites__recruiter=user_profile)
    
    # Ordenação por score
    score_sort = request.GET.get('score_sort', '')
    if score_sort == 'high_to_low':
        applications = applications.annotate(
            avg_score=(Avg('evaluations__technical_score') + 
                     Avg('evaluations__experience_score') + 
                     Avg('evaluations__cultural_fit_score')) / 3
        ).order_by('-avg_score')
    elif score_sort == 'low_to_high':
        applications = applications.annotate(
            avg_score=(Avg('evaluations__technical_score') + 
                     Avg('evaluations__experience_score') + 
                     Avg('evaluations__cultural_fit_score')) / 3
        ).order_by('avg_score')
    else:
        # Ordenação padrão por data de criação
        applications = applications.order_by('-created_at')
    
    # Adiciona informações sobre favoritos
    for application in applications:
        application.is_favorite = application.favorites.filter(recruiter=user_profile).exists()
        
        # Calcula o score médio se houver avaliações
        if application.evaluations.exists():
            evaluation = application.evaluations.first()
            total_score = evaluation.technical_score + evaluation.experience_score + evaluation.cultural_fit_score
            application.avg_score = round(total_score / 3, 1)
        else:
            application.avg_score = None
    
    # Contagem de candidaturas por status
    status_counts = {
        'total': applications.count(),
        'pending': applications.filter(status='pending').count(),
        'under_review': applications.filter(status='under_review').count(),
        'interview': applications.filter(status='interview').count(),
        'approved': applications.filter(status='approved').count(),
        'rejected': applications.filter(status='rejected').count(),
        'withdrawn': applications.filter(status='withdrawn').count(),
    }
    
    # Obtém dados para os filtros - todas as vagas e hospitais
    vacancies = Vacancy.objects.all()
    hospitals = Hospital.objects.all()
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(applications, 10)
    page_number = request.GET.get('page')
    applications = paginator.get_page(page_number)
    
    context = {
        'applications': applications,
        'vacancies': vacancies,
        'hospitals': hospitals,
        'status_counts': status_counts,
        'status_filter': status_filter,
        'vacancy_filter': vacancy_filter,
        'hospital_filter': hospital_filter,
        'date_filter': date_filter,
        'search_query': search_query,
        'score_sort': score_sort,
        'favorites_filter': favorites_filter,
        'page_title': _('Candidaturas'),
    }
    
    return render(request, 'applications/recruiter_application_list.html', context)


@login_required
def minhas_candidaturas(request):
    """
    View para exibir minhas candidaturas para candidatos.
    """
    # Verifica se o usuário é um candidato
    if request.user.role != 'candidate':
        messages.error(request, _('Apenas candidatos podem acessar esta página.'))
        return redirect('home')
    
    # Filtros
    status_filter = request.GET.get('status', '')
    vacancy_filter = request.GET.get('vacancy', '')
    date_filter = request.GET.get('date', '')
    
    # Função auxiliar para determinar cor do badge
    def get_status_badge_color(status):
        colors = {
            'pending': 'warning',
            'under_review': 'info',
            'interview': 'primary',
            'approved': 'success',
            'rejected': 'danger',
            'withdrawn': 'secondary'
        }
        return colors.get(status, 'secondary')
    
    # Obtém candidaturas reais do usuário
    try:
        user_profile = request.user.profile
        applications = Application.objects.filter(candidate=user_profile).select_related(
            'vacancy__hospital', 'vacancy__department'
        ).order_by('-created_at')
        
        # Aplica filtros ANTES de processar os dados
        if status_filter:
            applications = applications.filter(status=status_filter)
        
        if vacancy_filter:
            applications = applications.filter(id=vacancy_filter)
        
        if date_filter:
            from datetime import datetime, timedelta
            today = datetime.now().date()
            
            if date_filter == '7':  # Últimos 7 dias
                start_date = today - timedelta(days=7)
                applications = applications.filter(created_at__date__gte=start_date)
            elif date_filter == '30':  # Últimos 30 dias
                start_date = today - timedelta(days=30)
                applications = applications.filter(created_at__date__gte=start_date)
            elif date_filter == '90':  # Últimos 90 dias
                start_date = today - timedelta(days=90)
                applications = applications.filter(created_at__date__gte=start_date)
        

        
        # Prepara dados para o template
        candidaturas_data = []
        for app in applications:
            # Determina próxima etapa baseada no status
            proxima_etapa = "Aguardando análise do currículo"
            if app.status == 'under_review':
                proxima_etapa = "Em análise pelos recrutadores"
            elif app.status == 'interview':
                proxima_etapa = "Entrevista agendada"
            elif app.status == 'approved':
                proxima_etapa = "Aguardando contato do RH"
            elif app.status == 'rejected':
                proxima_etapa = "Candidatura encerrada"
            elif app.status == 'withdrawn':
                proxima_etapa = "Candidatura cancelada"
            
            candidaturas_data.append({
                'id': app.id,
                'vaga': app.vacancy.title,
                'unidade': f"{app.vacancy.hospital.name} - {app.vacancy.hospital.city}, {app.vacancy.hospital.state}",
                'data_candidatura': app.created_at.strftime('%d/%m/%Y'),
                'status': app.get_status_display(),
                'status_badge': f"bg-{get_status_badge_color(app.status)}",
                'proxima_etapa': proxima_etapa,
                'application': app,  # Para usar no template
            })
        
        # Simula dados de próximas entrevistas (pode ser implementado no futuro)
        entrevistas_data = []
        
        # Paginação
        from django.core.paginator import Paginator
        paginator = Paginator(candidaturas_data, 10)
        page_number = request.GET.get('page')
        candidaturas_data = paginator.get_page(page_number)
        
    except Exception as e:
        print(f"Erro ao carregar candidaturas: {e}")
        candidaturas_data = []
        entrevistas_data = []
    
    # Obtém dados para filtros
    hospitals = Hospital.objects.all()
    
    context = {
        'candidaturas': candidaturas_data,
        'entrevistas': entrevistas_data,
        'hospitals': hospitals,
        'status_filter': status_filter,
        'vacancy_filter': vacancy_filter,
        'date_filter': date_filter,
        'page_title': _('Minhas Candidaturas'),
    }
    
    return render(request, 'applications/minhas_candidaturas.html', context)


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def toggle_favorite(request, application_id):
    """
    Adiciona ou remove uma candidatura dos favoritos.
    """
    if request.user.role not in ['recruiter', 'recrutador', 'admin']:
        return JsonResponse({'error': 'Permissão negada'}, status=403)
    
    try:
        application = Application.objects.get(pk=application_id)
        user_profile = request.user.profile
        
        # Verifica se já é favorito
        try:
            favorite = ApplicationFavorite.objects.get(
                application=application,
                recruiter=user_profile
            )
            is_favorite = True
        except ApplicationFavorite.DoesNotExist:
            favorite = None
            is_favorite = False
        
        # Processa a ação baseada no estado atual
        if is_favorite:
            # Remove dos favoritos
            favorite.delete()
            is_favorite = False
            message = 'Candidatura removida dos favoritos'
        else:
            # Adiciona aos favoritos
            ApplicationFavorite.objects.create(
                application=application,
                recruiter=user_profile
            )
            is_favorite = True
            message = 'Candidatura adicionada aos favoritos'
        
        return JsonResponse({
            'success': True,
            'is_favorite': is_favorite,
            'message': message
        })
        
    except Application.DoesNotExist:
        return JsonResponse({'error': 'Candidatura não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def export_candidaturas(request):
    """
    Exporta candidaturas filtradas para CSV ou Excel.
    """
    # Verifica se o usuário é um recrutador
    if request.user.role not in ['recruiter', 'recrutador', 'admin']:
        messages.error(request, _('Apenas recrutadores podem acessar esta página.'))
        return redirect('home')
    
    # Obtém o formato de exportação
    export_format = request.GET.get('format', 'csv')
    
    # Aplica os mesmos filtros da view de candidaturas
    applications = Application.objects.all()
    
    # Filtros
    status_filter = request.GET.get('status', '')
    if status_filter:
        applications = applications.filter(status=status_filter)
    
    vacancy_filter = request.GET.get('vacancy', '')
    if vacancy_filter:
        applications = applications.filter(vacancy_id=vacancy_filter)
    
    hospital_filter = request.GET.get('hospital', '')
    if hospital_filter:
        applications = applications.filter(vacancy__hospital_id=hospital_filter)
    
    date_filter = request.GET.get('date', '')
    if date_filter:
        today = timezone.now().date()
        if date_filter == 'today':
            applications = applications.filter(created_at__date=today)
        elif date_filter == 'week':
            week_ago = today - timedelta(days=7)
            applications = applications.filter(created_at__date__gte=week_ago)
        elif date_filter == 'month':
            month_ago = today - timedelta(days=30)
            applications = applications.filter(created_at__date__gte=month_ago)
    
    search_filter = request.GET.get('search', '')
    if search_filter:
        applications = applications.filter(
            Q(candidate__user__first_name__icontains=search_filter) |
            Q(candidate__user__last_name__icontains=search_filter) |
            Q(candidate__user__email__icontains=search_filter)
        )
    
    # Filtro de favoritos
    favorites_filter = request.GET.get('favorites', '')
    if favorites_filter == 'true':
        applications = applications.filter(favorites__recruiter=request.user.profile)
    
    # Ordenação por score
    score_sort = request.GET.get('score_sort', '')
    if score_sort == 'high_to_low':
        applications = applications.annotate(
            avg_score=(Avg('evaluations__technical_score') + 
                     Avg('evaluations__experience_score') + 
                     Avg('evaluations__cultural_fit_score')) / 3
        ).order_by('-avg_score')
    elif score_sort == 'low_to_high':
        applications = applications.annotate(
            avg_score=(Avg('evaluations__technical_score') + 
                     Avg('evaluations__experience_score') + 
                     Avg('evaluations__cultural_fit_score')) / 3
        ).order_by('avg_score')
    else:
        # Ordenação padrão por data de criação
        applications = applications.order_by('-created_at')
    
    # Prepara os dados para exportação
    data = []
    for application in applications:
        # Calcula o score médio se houver avaliações
        avg_score = None
        if application.evaluations.exists():
            evaluation = application.evaluations.first()
            total_score = evaluation.technical_score + evaluation.experience_score + evaluation.cultural_fit_score
            avg_score = round(total_score / 3, 1)
        
        # Obtém informações do candidato
        candidate = application.candidate
        user = candidate.user
        
        # Tenta obter o perfil de candidato se existir
        candidate_profile = None
        try:
            candidate_profile = user.candidate_profile
        except:
            pass
        
        # Obtém informações da vaga e hospital
        vacancy = application.vacancy
        hospital_name = vacancy.hospital.name if vacancy.hospital else "N/A"
        hospital_location = f"{vacancy.hospital.city}, {vacancy.hospital.state}" if vacancy.hospital else "N/A"
        
        # Status em português
        status_map = {
            'pending': 'Pendente',
            'under_review': 'Em Análise',
            'interview': 'Entrevista',
            'approved': 'Aprovado',
            'rejected': 'Rejeitado',
            'withdrawn': 'Desistência'
        }
        
        # Obtém dados priorizando CandidateProfile quando disponível
        telefone = user.phone
        if candidate_profile and candidate_profile.phone:
            telefone = candidate_profile.phone
        
        whatsapp = user.whatsapp
        if candidate_profile and candidate_profile.whatsapp:
            whatsapp = candidate_profile.whatsapp
            
        cpf = user.cpf
        if candidate_profile and candidate_profile.cpf:
            cpf = candidate_profile.cpf
            
        endereco = user.address
        if candidate_profile and candidate_profile.address:
            endereco = candidate_profile.address
            
        cidade = user.city
        if candidate_profile and candidate_profile.city:
            cidade = candidate_profile.city
            
        estado = user.state
        if candidate_profile and candidate_profile.state:
            estado = candidate_profile.state
            
        cep = user.zip_code
        if candidate_profile and candidate_profile.zip_code:
            cep = candidate_profile.zip_code
        
        row = {
            'Nome Completo': user.get_full_name(),
            'Email': user.email,
            'Telefone': telefone or 'N/A',
            'WhatsApp': whatsapp or 'N/A',
            'CPF': cpf or 'N/A',
            'Data de Nascimento': user.date_of_birth.strftime('%d/%m/%Y') if user.date_of_birth else 'N/A',
            'Endereço': endereco or 'N/A',
            'Cidade': cidade or 'N/A',
            'Estado': estado or 'N/A',
            'CEP': cep or 'N/A',
            'Vaga': vacancy.title,
            'Hospital': hospital_name,
            'Localização': hospital_location,
            'Status': status_map.get(application.status, application.status),
            'Score': avg_score if avg_score else 'N/A',
            'Data da Candidatura': application.created_at.strftime('%d/%m/%Y %H:%M'),
            'Carta de Apresentação': application.cover_letter or 'N/A',
            'Notas do Recrutador': application.recruiter_notes or 'N/A'
        }
        data.append(row)
    
    # Nome do arquivo
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'candidaturas_{timestamp}'
    
    if export_format == 'excel':
        return export_as_excel_candidaturas(data, filename)
    else:
        return export_as_csv_candidaturas(data, filename)


def export_as_csv_candidaturas(data, filename):
    """
    Exporta candidaturas para CSV.
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    # Adiciona BOM para UTF-8
    response.write('\ufeff')
    
    if not data:
        return response
    
    writer = csv.DictWriter(response, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    
    return response


def export_as_excel_candidaturas(data, filename):
    """
    Exporta candidaturas para Excel.
    """
    if not data:
        # Retorna um arquivo Excel vazio
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidaturas"
        ws['A1'] = "Nenhuma candidatura encontrada"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb.save(response)
        return response
    
    # Cria um DataFrame com os dados
    df = pd.DataFrame(data)
    
    # Cria o arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidaturas"
    
    # Adiciona os dados
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Ajusta a largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Formata o cabeçalho
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb.save(response)
    return response


# API Views

class ApplicationViewSet(viewsets.ModelViewSet):
    """
    API endpoint para candidaturas.
    """
    queryset = Application.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrRecruiter]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['vacancy__title', 'candidate__user__first_name', 'candidate__user__last_name']
    ordering_fields = ['created_at', 'updated_at', 'status']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ApplicationCreateSerializer
        elif self.action == 'update_status':
            return ApplicationStatusUpdateSerializer
        return ApplicationSerializer
    
    def get_queryset(self):
        user_profile = self.request.user.profile
        
        # Candidatos veem apenas suas próprias candidaturas
        if request.user.role == 'candidate':
            return Application.objects.filter(candidate=user_profile)
        
        # Recrutadores e administradores veem todas as candidaturas
        return Application.objects.all()
    
    def perform_create(self, serializer):
        serializer.save(candidate=self.request.user.profile)
    
    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        """
        Atualiza o status de uma candidatura.
        """
        application = self.get_object()
        serializer = self.get_serializer(application, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ApplicationEvaluationViewSet(viewsets.ModelViewSet):
    """
    API endpoint para avaliações de candidaturas.
    """
    queryset = ApplicationEvaluation.objects.all()
    serializer_class = ApplicationEvaluationSerializer
    permission_classes = [permissions.IsAuthenticated, IsRecruiterOrAdmin]
    
    def get_queryset(self):
        user_profile = self.request.user.profile
        
        # Recrutadores veem apenas suas próprias avaliações
        if request.user.role == 'recruiter':
            return ApplicationEvaluation.objects.filter(evaluator=user_profile)
        
        # Administradores veem todas as avaliações
        return ApplicationEvaluation.objects.all()
    
    def perform_create(self, serializer):
        application_id = self.request.data.get('application')
        application = get_object_or_404(Application, pk=application_id)
        
        serializer.save(
            evaluator=self.request.user.profile,
            application=application
        )


class ResumeViewSet(viewsets.ModelViewSet):
    """
    API endpoint para currículos.
    """
    queryset = Resume.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsResumeOwner]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ResumeCreateUpdateSerializer
        return ResumeSerializer
    
    def get_queryset(self):
        user_profile = self.request.user.profile
        
        # Candidatos veem apenas seu próprio currículo
        if request.user.role == 'candidate':
            return Resume.objects.filter(candidate=user_profile)
        
        # Recrutadores e administradores veem todos os currículos
        return Resume.objects.all()
    
    def perform_create(self, serializer):
        serializer.save(candidate=self.request.user.profile)


class EducationViewSet(viewsets.ModelViewSet):
    """
    API endpoint para formação educacional.
    """
    queryset = Education.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsEducationOrExperienceOwner]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return EducationCreateUpdateSerializer
        return EducationSerializer
    
    def get_queryset(self):
        user_profile = self.request.user.profile
        
        # Candidatos veem apenas suas próprias formações educacionais
        if request.user.role == 'candidate':
            return Education.objects.filter(resume__candidate=user_profile)
        
        # Recrutadores e administradores veem todas as formações educacionais
        return Education.objects.all()
    
    def perform_create(self, serializer):
        resume_id = self.request.data.get('resume_id')
        resume = get_object_or_404(Resume, pk=resume_id)
        
        # Verifica se o usuário é o proprietário do currículo
        if resume.candidate != self.request.user.profile:
            raise permissions.PermissionDenied(_('Você não tem permissão para adicionar formação a este currículo.'))
        
        serializer.save(resume=resume)


class WorkExperienceViewSet(viewsets.ModelViewSet):
    """
    API endpoint para experiências profissionais.
    """
    queryset = WorkExperience.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsEducationOrExperienceOwner]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return WorkExperienceCreateUpdateSerializer
        return WorkExperienceSerializer
    
    def get_queryset(self):
        user_profile = self.request.user.profile
        
        # Candidatos veem apenas suas próprias experiências profissionais
        if request.user.role == 'candidate':
            return WorkExperience.objects.filter(resume__candidate=user_profile)
        
        # Recrutadores e administradores veem todas as experiências profissionais
        return WorkExperience.objects.all()
    
    def perform_create(self, serializer):
        resume_id = self.request.data.get('resume_id')
        resume = get_object_or_404(Resume, pk=resume_id)
        
        # Verifica se o usuário é o proprietário do currículo
        if resume.candidate != self.request.user.profile:
            raise permissions.PermissionDenied(_('Você não tem permissão para adicionar experiência a este currículo.'))
        
        serializer.save(resume=resume)


@api_view(['GET'])
def available_for_interview(request):
    """
    API endpoint para retornar candidaturas disponíveis para agendar entrevista.
    Disponível apenas para recrutadores e administradores.
    """
    # Debug info
    print(f"DEBUG: User authenticated: {request.user.is_authenticated}")
    print(f"DEBUG: User: {request.user}")
    if hasattr(request.user, 'role'):
        print(f"DEBUG: User role: {request.user.role}")
    
    # Verificar se o usuário tem permissão
    if not request.user.is_authenticated or request.user.role not in ['recruiter', 'admin']:
        return Response({'error': 'Acesso negado', 'debug': {'authenticated': request.user.is_authenticated, 'role': getattr(request.user, 'role', 'N/A')}}, status=status.HTTP_403_FORBIDDEN)
    
    # Filtrar candidaturas baseado no tipo de usuário
    if request.user.role == 'recruiter':
        # Recrutadores veem apenas candidaturas de suas vagas
        applications = Application.objects.filter(
            vacancy__recruiter=request.user,
            status__in=['pending', 'submitted', 'under_review', 'reviewed', 'interview_scheduled']
        ).select_related('candidate', 'vacancy')
    else:
        # Administradores veem todas as candidaturas
        applications = Application.objects.filter(
            status__in=['pending', 'submitted', 'under_review', 'reviewed', 'interview_scheduled']
        ).select_related('candidate', 'vacancy')
    
    # Serializar os dados
    data = []
    for app in applications:
        candidate_name = 'N/A'
        candidate_email = 'N/A'
        
        if app.candidate:
            candidate_name = app.candidate.user.get_full_name()
            candidate_email = app.candidate.user.email
        
        data.append({
            'id': app.id,
            'candidate_name': candidate_name,
            'candidate_email': candidate_email,
            'vacancy_title': app.vacancy.title,
            'vacancy_id': app.vacancy.id,
            'status': app.get_status_display(),
            'applied_at': app.created_at.strftime('%d/%m/%Y %H:%M')
        })
    
    return Response(data)
